"""
ChurchOS Reports Module
Generates PDF and Excel reports for all major modules.
"""
import io
from datetime import date
from decimal import Decimal
from django.utils import timezone
from django.http import HttpResponse, FileResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status


def _excel_response(wb, filename):
    """Helper to return an Excel file response."""
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


def _get_date_range(request):
    today = timezone.now().date()
    date_from = request.query_params.get('date_from', today.replace(day=1).isoformat())
    date_to = request.query_params.get('date_to', today.isoformat())
    fmt = request.query_params.get('format', 'json')
    return date_from, date_to, fmt


class FinanceReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.finance.models import Transaction, TransactionType
        from django.db.models import Sum

        date_from, date_to, fmt = _get_date_range(request)
        qs = Transaction.objects.filter(
            is_active=True,
            transaction_date__gte=date_from,
            transaction_date__lte=date_to,
        ).select_related('member', 'category', 'recorded_by').order_by('transaction_date')

        income_types = [
            TransactionType.TITHE, TransactionType.OFFERING,
            TransactionType.DONATION, TransactionType.PLEDGE_PAYMENT,
            TransactionType.HALL_RENTAL, TransactionType.OTHER_INCOME,
        ]
        income_qs = qs.filter(transaction_type__in=income_types)
        expense_qs = qs.filter(transaction_type=TransactionType.EXPENSE)

        if fmt == 'excel':
            return self._excel(qs, date_from, date_to, income_qs, expense_qs)

        # JSON summary
        total_income = income_qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')
        total_expenses = expense_qs.aggregate(t=Sum('amount'))['t'] or Decimal('0')
        by_type = {}
        for row in qs.values('transaction_type').annotate(t=Sum('amount')):
            by_type[row['transaction_type']] = float(row['t'])

        return Response({
            'period': {'from': date_from, 'to': date_to},
            'total_income': float(total_income),
            'total_expenses': float(total_expenses),
            'net': float(total_income - total_expenses),
            'by_type': by_type,
            'transactions': [
                {
                    'date': str(t.transaction_date),
                    'reference': t.reference,
                    'receipt': t.receipt_number,
                    'type': t.transaction_type,
                    'member': t.member.get_full_name() if t.member else 'Anonymous',
                    'amount': float(t.amount),
                    'payment_method': t.payment_method,
                    'verified': t.verified,
                }
                for t in qs[:500]
            ]
        })

    def _excel(self, qs, date_from, date_to, income_qs, expense_qs):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from django.db.models import Sum

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Finance Report'

            # Header
            ws.merge_cells('A1:H1')
            ws['A1'] = f'Finance Report: {date_from} to {date_to}'
            ws['A1'].font = Font(bold=True, size=14)

            # Column headers
            headers = ['Date', 'Reference', 'Receipt No', 'Type', 'Member', 'Amount', 'Method', 'Verified']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor='1a6b3c')
                cell.alignment = Alignment(horizontal='center')

            # Data
            for row_idx, t in enumerate(qs[:1000], 4):
                ws.cell(row=row_idx, column=1, value=str(t.transaction_date))
                ws.cell(row=row_idx, column=2, value=t.reference)
                ws.cell(row=row_idx, column=3, value=t.receipt_number)
                ws.cell(row=row_idx, column=4, value=t.transaction_type)
                ws.cell(row=row_idx, column=5, value=t.member.get_full_name() if t.member else 'Anonymous')
                ws.cell(row=row_idx, column=6, value=float(t.amount))
                ws.cell(row=row_idx, column=7, value=t.payment_method)
                ws.cell(row=row_idx, column=8, value='Yes' if t.verified else 'No')

            # Summary sheet
            ws2 = wb.create_sheet('Summary')
            total_inc = income_qs.aggregate(t=Sum('amount'))['t'] or 0
            total_exp = expense_qs.aggregate(t=Sum('amount'))['t'] or 0
            ws2['A1'] = 'Summary'
            ws2['A1'].font = Font(bold=True, size=12)
            ws2['A3'] = 'Total Income'
            ws2['B3'] = float(total_inc)
            ws2['A4'] = 'Total Expenses'
            ws2['B4'] = float(total_exp)
            ws2['A5'] = 'Net Balance'
            ws2['B5'] = float(total_inc - total_exp)
            ws2['B5'].font = Font(bold=True)

            # Auto-width
            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

            return _excel_response(wb, f'finance_report_{date_from}_{date_to}.xlsx')
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=500)


class MembersReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.members.models import Member, MemberStatus

        _, _, fmt = _get_date_range(request)
        membership_status = request.query_params.get('membership_status', '')
        qs = Member.objects.filter(is_active=True).select_related('zone', 'cell_group').order_by('last_name', 'first_name')
        if membership_status:
            qs = qs.filter(membership_status=membership_status)

        if fmt == 'excel':
            return self._excel(qs)

        return Response({
            'count': qs.count(),
            'members': [
                {
                    'member_number': m.member_number,
                    'name': m.get_full_name(),
                    'gender': m.gender,
                    'date_of_birth': str(m.date_of_birth) if m.date_of_birth else '',
                    'phone': m.phone_primary,
                    'email': m.email,
                    'status': m.membership_status,
                    'zone': m.zone.name if m.zone else '',
                    'cell_group': m.cell_group.name if m.cell_group else '',
                    'membership_date': str(m.membership_date) if m.membership_date else '',
                }
                for m in qs[:1000]
            ]
        })

    def _excel(self, qs):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Members'

            headers = ['No.', 'Member No', 'Full Name', 'Gender', 'DOB', 'Phone', 'Email', 'Status', 'Zone', 'Cell Group', 'Member Since']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor='1a6b3c')

            for row_idx, m in enumerate(qs[:2000], 2):
                ws.cell(row=row_idx, column=1, value=row_idx - 1)
                ws.cell(row=row_idx, column=2, value=m.member_number)
                ws.cell(row=row_idx, column=3, value=m.get_full_name())
                ws.cell(row=row_idx, column=4, value=m.gender)
                ws.cell(row=row_idx, column=5, value=str(m.date_of_birth) if m.date_of_birth else '')
                ws.cell(row=row_idx, column=6, value=m.phone_primary)
                ws.cell(row=row_idx, column=7, value=m.email)
                ws.cell(row=row_idx, column=8, value=m.membership_status)
                ws.cell(row=row_idx, column=9, value=m.zone.name if m.zone else '')
                ws.cell(row=row_idx, column=10, value=m.cell_group.name if m.cell_group else '')
                ws.cell(row=row_idx, column=11, value=str(m.membership_date) if m.membership_date else '')

            for col in ws.columns:
                max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

            return _excel_response(wb, 'members_report.xlsx')
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=500)


class AttendanceReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.events.models import EventAttendance
        from django.db.models import Count

        date_from, date_to, fmt = _get_date_range(request)
        qs = EventAttendance.objects.filter(
            date__gte=date_from, date__lte=date_to
        ).select_related('member', 'service_type').order_by('-date')

        if fmt == 'excel':
            return self._excel(qs, date_from, date_to)

        by_date = {}
        for a in qs:
            key = str(a.date)
            by_date[key] = by_date.get(key, 0) + 1

        return Response({
            'period': {'from': date_from, 'to': date_to},
            'total_records': qs.count(),
            'by_date': by_date,
            'records': [
                {
                    'date': str(a.date),
                    'member': a.member.get_full_name() if a.member else a.visitor_name or 'Visitor',
                    'service_type': a.service_type.name if a.service_type else '',
                    'is_visitor': a.is_visitor,
                }
                for a in qs[:500]
            ]
        })

    def _excel(self, qs, date_from, date_to):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Attendance'
            headers = ['Date', 'Member/Visitor', 'Service Type', 'Is Visitor']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor='1a6b3c')
            for row_idx, a in enumerate(qs[:2000], 2):
                ws.cell(row=row_idx, column=1, value=str(a.date))
                ws.cell(row=row_idx, column=2, value=a.member.get_full_name() if a.member else a.visitor_name or 'Visitor')
                ws.cell(row=row_idx, column=3, value=a.service_type.name if a.service_type else '')
                ws.cell(row=row_idx, column=4, value='Yes' if a.is_visitor else 'No')
            return _excel_response(wb, f'attendance_{date_from}_{date_to}.xlsx')
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=500)


class PayrollReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.workers.models import Payslip, PayrollRun

        _, _, fmt = _get_date_range(request)
        month = request.query_params.get('month', timezone.now().month)
        year = request.query_params.get('year', timezone.now().year)
        qs = Payslip.objects.filter(
            payroll_run__month=month, payroll_run__year=year
        ).select_related('worker', 'payroll_run').order_by('worker__last_name')

        if fmt == 'excel':
            return self._excel(qs, month, year)

        from django.db.models import Sum
        totals = qs.aggregate(gross=Sum('gross_salary'), deductions=Sum('total_deductions'), net=Sum('net_salary'))
        return Response({
            'period': f'{month}/{year}',
            'count': qs.count(),
            'total_gross': float(totals['gross'] or 0),
            'total_deductions': float(totals['deductions'] or 0),
            'total_net': float(totals['net'] or 0),
            'payslips': [
                {
                    'employee_id': p.worker.employee_id,
                    'name': p.worker.get_full_name(),
                    'basic': float(p.basic_salary),
                    'allowances': float(p.total_allowances),
                    'gross': float(p.gross_salary),
                    'deductions': float(p.total_deductions),
                    'net': float(p.net_salary),
                    'status': p.payment_status,
                }
                for p in qs
            ]
        })

    def _excel(self, qs, month, year):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Payroll'
            headers = ['EMP ID', 'Name', 'Basic Salary', 'Allowances', 'Gross', 'Deductions', 'Net Salary', 'Status']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor='1a6b3c')
            for row_idx, p in enumerate(qs, 2):
                ws.cell(row=row_idx, column=1, value=p.worker.employee_id)
                ws.cell(row=row_idx, column=2, value=p.worker.get_full_name())
                ws.cell(row=row_idx, column=3, value=float(p.basic_salary))
                ws.cell(row=row_idx, column=4, value=float(p.total_allowances))
                ws.cell(row=row_idx, column=5, value=float(p.gross_salary))
                ws.cell(row=row_idx, column=6, value=float(p.total_deductions))
                ws.cell(row=row_idx, column=7, value=float(p.net_salary))
                ws.cell(row=row_idx, column=8, value=p.payment_status)
            return _excel_response(wb, f'payroll_{month}_{year}.xlsx')
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=500)


class TitheReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.finance.models import Transaction, TransactionType
        from django.db.models import Sum, Count

        date_from, date_to, fmt = _get_date_range(request)
        qs = (
            Transaction.objects
            .filter(
                transaction_type=TransactionType.TITHE,
                transaction_date__gte=date_from,
                transaction_date__lte=date_to,
                is_active=True,
            )
            .values('member__id', 'member__first_name', 'member__last_name', 'member__member_number')
            .annotate(total=Sum('amount'), payments=Count('id'))
            .order_by('-total')
        )

        if fmt == 'excel':
            return self._excel(list(qs), date_from, date_to)

        return Response({
            'period': {'from': date_from, 'to': date_to},
            'records': [
                {
                    'member_number': r['member__member_number'],
                    'name': f"{r['member__first_name']} {r['member__last_name']}".strip(),
                    'total': float(r['total']),
                    'payments': r['payments'],
                }
                for r in qs
            ]
        })

    def _excel(self, data, date_from, date_to):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Tithe Report'
            headers = ['Member No', 'Name', 'Total Tithe', 'No. of Payments']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor='1a6b3c')
            for row_idx, r in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=r['member__member_number'])
                ws.cell(row=row_idx, column=2, value=f"{r['member__first_name']} {r['member__last_name']}".strip())
                ws.cell(row=row_idx, column=3, value=float(r['total']))
                ws.cell(row=row_idx, column=4, value=r['payments'])
            return _excel_response(wb, f'tithe_report_{date_from}_{date_to}.xlsx')
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=500)


class InventoryReportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.inventory.models import InventoryItem
        from django.db.models import Sum

        _, _, fmt = _get_date_range(request)
        qs = InventoryItem.objects.filter(is_active=True).select_related('category', 'custodian').order_by('category__name', 'name')

        if fmt == 'excel':
            return self._excel(qs)

        total_value = qs.aggregate(t=Sum('current_value'))['t'] or 0
        return Response({
            'count': qs.count(),
            'total_value': float(total_value),
            'items': [
                {
                    'name': i.name,
                    'category': i.category.name if i.category else '',
                    'serial_number': i.serial_number,
                    'condition': i.condition,
                    'quantity': i.quantity,
                    'location': i.location,
                    'purchase_price': float(i.purchase_price) if i.purchase_price else None,
                    'current_value': float(i.current_value) if i.current_value else None,
                    'custodian': i.custodian.get_full_name() if i.custodian else '',
                }
                for i in qs
            ]
        })

    def _excel(self, qs):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Inventory'
            headers = ['Name', 'Category', 'Serial No', 'Condition', 'Qty', 'Location', 'Department', 'Purchase Price', 'Current Value', 'Custodian']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor='1a6b3c')
            for row_idx, i in enumerate(qs, 2):
                ws.cell(row=row_idx, column=1, value=i.name)
                ws.cell(row=row_idx, column=2, value=i.category.name if i.category else '')
                ws.cell(row=row_idx, column=3, value=i.serial_number)
                ws.cell(row=row_idx, column=4, value=i.condition)
                ws.cell(row=row_idx, column=5, value=i.quantity)
                ws.cell(row=row_idx, column=6, value=i.location)
                ws.cell(row=row_idx, column=7, value=i.department)
                ws.cell(row=row_idx, column=8, value=float(i.purchase_price) if i.purchase_price else '')
                ws.cell(row=row_idx, column=9, value=float(i.current_value) if i.current_value else '')
                ws.cell(row=row_idx, column=10, value=i.custodian.get_full_name() if i.custodian else '')
            return _excel_response(wb, 'inventory_report.xlsx')
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=500)


class AuditReportExportView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.audit.models import AuditReport

        _, _, fmt = _get_date_range(request)
        qs = AuditReport.objects.select_related('check').order_by('-run_at')[:200]

        if fmt == 'excel':
            return self._excel(list(qs))

        return Response({
            'reports': [
                {
                    'check': r.check.name,
                    'category': r.check.category,
                    'severity': r.check.severity,
                    'status': r.status,
                    'summary': r.result_summary,
                    'affected': r.affected_count,
                    'run_at': str(r.run_at),
                    'resolved': r.resolved,
                }
                for r in qs
            ]
        })

    def _excel(self, data):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Audit Reports'
            headers = ['Check Name', 'Category', 'Severity', 'Status', 'Summary', 'Affected Count', 'Run At', 'Resolved']
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(fill_type='solid', fgColor='1a6b3c')
            for row_idx, r in enumerate(data, 2):
                ws.cell(row=row_idx, column=1, value=r.check.name)
                ws.cell(row=row_idx, column=2, value=r.check.category)
                ws.cell(row=row_idx, column=3, value=r.check.severity)
                ws.cell(row=row_idx, column=4, value=r.status)
                ws.cell(row=row_idx, column=5, value=r.result_summary)
                ws.cell(row=row_idx, column=6, value=r.affected_count)
                ws.cell(row=row_idx, column=7, value=str(r.run_at))
                ws.cell(row=row_idx, column=8, value='Yes' if r.resolved else 'No')
            return _excel_response(wb, 'audit_report.xlsx')
        except ImportError:
            return Response({'error': 'openpyxl not installed'}, status=500)
